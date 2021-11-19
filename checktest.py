import pandas as pd
from numpy import nan
from defaultfigure import *
from initialize import initialize_dir_year, initialize_dir_region, get_cities, find_typedata
import os
import shutil
from shutil import copyfile
from openpyxl import load_workbook
import string
import random


def cancel_upd():
    year = initialize_dir_year()
    for y in year:
        directory_contents = os.listdir('SCBAA/'+y)
        init_list = list(directory_contents)
        for i in init_list:
            if(i.find("-copy.xlsx") != -1):
                os.remove('SCBAA/'+y+'/'+i)
    return None


def get_adminaccess():
    filepath = "SCBAAsamp/passw/passw.xlsx"
    reg_excel = pd.ExcelFile(filepath)
    city_excel = pd.read_excel(reg_excel, "Password")
    x = city_excel.iloc[0, 0]
    y = city_excel.iloc[1, 0]
    z = city_excel.iloc[2, 0]
    return x, y, z


def id_generator(size=15, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def confirm_upd():
    year = initialize_dir_year()
    for y in year:
        directory_contents = os.listdir('SCBAA/'+y)
        init_list = list(directory_contents)
        for i in init_list:
            del_file = ''
            if(i.find("-copy.xlsx") != -1):
                del_file = i.replace('-copy.xlsx', '')
                os.remove('SCBAA/'+y+'/'+del_file+'.xlsx')
                os.rename('SCBAA/'+y+'/'+del_file+'-copy.xlsx',
                          'SCBAA/'+y+'/'+del_file+'.xlsx')
                wb = load_workbook('SCBAA/'+y+'/'+del_file+'.xlsx')
                wb.save('SCBAA/'+y+'/'+del_file+'.xlsx')
    return None


def update_scan():
    year = initialize_dir_year()
    lst_val = [9, 10, 11, 14, 15, 16, 19, 20, 22, 23, 24, 25, 27, 28, 29, 31, 32, 33, 34, 40, 41, 42, 44, 45, 46, 48, 49, 50, 52, 53, 54, 56,
               57, 58, 60, 61, 62, 64, 65, 66, 68, 69, 70, 73, 74, 76, 77, 79, 80, 82, 83, 85, 86, 88, 89, 90, 94, 96, 98, 100, 102, 104, 106, 108]
    lst_checkerrs = {"path": [], "fix": []}
    total_amlst = [12, 17, 35, 91]
    region = initialize_dir_region()['value']
    lst_checkmissingscbaa = []
    dict_errs = {}
    for y in year:
        y = int(y)
        dict_errs[y] = {}
        for r in region:
            filepath = "SCBAA/" + str(y) + "/" + r + ".xlsx"
            reg_excel = pd.ExcelFile(filepath)
            for c in dict_scbaa["Region"][r]:
                app, rev = 0, 0
                city_excel = pd.read_excel(reg_excel, c)
                rev = city_excel.iloc[35, 4]
                app = city_excel.iloc[110, 4]
                cityerrlst = []
                city = {}
                for i in range(102):
                    val = city_excel.iloc[i+7, 4]
                    if i+7 in lst_val:
                        if str(val) == "nan":
                            lst_checkerrs['path'].append(
                                "SCBAA/"+str(y)+"/"+r+".xlsx/"+c+"/E"+str((i+7)+2))
                            lst_checkerrs['fix'].append(
                                str(val) + " -> float value")
                            cityerrlst.append(i)
                    elif i+7 in total_amlst:
                        pass
                    else:
                        if str(val) != "nan":
                            lst_checkerrs['path'].append(
                                "SCBAA/"+str(y)+"/"+r+".xlsx/"+c+"/E"+str((i+7)+2))
                            lst_checkerrs['fix'].append(str(val) + " -> nan")
                            cityerrlst.append(i)
                if rev == 0 and app == 0:
                    lst_checkmissingscbaa.append(
                        "SCBAA/"+str(y)+"/"+r+".xlsx/"+c)
                if cityerrlst:
                    city = {c: cityerrlst}
                    try:
                        dict_errs[y][r].update(city)
                    except KeyError:
                        dict_errs[y] = {r: city}
    print(dict_errs)
    for yr in dict_errs:
        for rg in dict_errs[yr]:
            filepath = "SCBAA/" + str(yr) + "/" + rg + ".xlsx"
            copyfile(filepath, "SCBAA/" + str(yr) + "/" + rg + "-copy.xlsx")
            wb = load_workbook("SCBAA/" + str(yr) + "/" + rg + "-copy.xlsx")
            for ct in dict_errs[yr][rg]:
                ws = wb[ct]
                for error in dict_errs[yr][rg][ct]:
                    if error+7 in lst_val:
                        ws["E"+str((error+7)+2)] = 0.0
                    elif error+7 in total_amlst:
                        pass
                    else:
                        ws["E"+str((error+7)+2)] = None
                try:
                    ws["E14"] = float(ws["E11"].value) + \
                        float(ws["E12"].value) + float(ws["E13"].value)
                    ws["E19"] = float(ws["E16"].value) + \
                        float(ws["E17"].value) + float(ws["E18"].value)
                    ws["E37"] = float(ws["E14"].value) + float(ws["E19"].value) + float(ws["E21"].value) + float(ws["E22"].value) + float(ws["E24"].value) + float(ws["E25"].value) + float(ws["E26"].value) + \
                        float(ws["E27"].value) + float(ws["E29"].value) + float(ws["E30"].value) + float(ws["E31"].value) + \
                        float(ws["E33"].value) + float(ws["E34"].value) + \
                        float(ws["E35"].value) + float(ws["E36"].value)
                    ws["E93"] = float(ws["E42"].value) + float(ws["E43"].value) + float(ws["E44"].value) + float(ws["E46"].value) + float(ws["E47"].value) + float(ws["E48"].value) + float(ws["E50"].value) + float(ws["E51"].value) + float(ws["E52"].value) + \
                        float(ws["E54"].value) + float(ws["E55"].value) + float(ws["E56"].value) + float(ws["E58"].value) + \
                        float(ws["E59"].value) + float(ws["E60"].value) + \
                        float(ws["E62"].value) + float(ws["E63"].value) + float(ws["E64"].value) + float(ws["E66"].value) + float(ws["E67"].value) + float(ws["E68"].value) + float(ws["E70"].value) + float(ws["E71"].value) + float(ws["E72"].value) + \
                        float(ws["E75"].value) + float(ws["E76"].value) + float(ws["E78"].value) + float(ws["E79"].value) + \
                        float(ws["E81"].value) + float(ws["E82"].value) + \
                        float(ws["E84"].value) + float(ws["E85"].value) + float(ws["E87"].value) + \
                        float(ws["E88"].value) + float(ws["E90"].value) + \
                        float(ws["E91"].value) + float(ws["E92"].value)
                    ws["E111"] = float(ws["E96"].value) + float(ws["E98"].value) + float(ws["E100"].value) + float(ws["E102"].value) + \
                        float(ws["E104"].value) + float(ws["E106"].value) + \
                        float(ws["E108"].value) + float(ws["E110"].value)
                    ws["E112"] = float(ws["E93"].value) + \
                        float(ws["E111"].value)
                except ValueError:
                    print(yr, rg, ct, "COPYPASTE")
                except TypeError:
                    print(yr, rg, ct, "FORMAT")
            wb.save("SCBAA/" + str(yr) + "/" + rg + "-copy.xlsx")
    return lst_checkmissingscbaa, lst_checkerrs


def update_defaultgraph():
    year = initialize_dir_year()
    region = initialize_dir_region()['value']
    dict_revapp = {"Region": [], "Year": [],
                   "Revenue": [], "Appropriations": []}
    for y in year:
        y = int(y)
        for r in region:
            reg_excel = pd.ExcelFile("SCBAA/" + str(y) + "/" + r + ".xlsx")
            totalrev = 0
            totalapp = 0
            for c in dict_scbaa["Region"][r]:
                app, rev = 0, 0
                city_excel = pd.read_excel(reg_excel, c)
                rev = city_excel.iloc[35, 4]
                app = city_excel.iloc[110, 4]
                totalrev += rev
                totalapp += app
            dict_revapp["Region"].append(r)
            dict_revapp["Year"].append(y)
            dict_revapp["Revenue"].append(totalrev)
            dict_revapp["Appropriations"].append(totalapp)
    df = pd.DataFrame(dict_revapp)
    df.to_excel("SCBAA/Defaultgraph.xlsx")
    return None


def delete_year(year):
    year = str(year)
    path = "SCBAA/"+year
    if(os.path.exists(path)):
        shutil.rmtree(path)
    else:
        pass
    return None


def add_year(year):
    year = str(year)
    path = "SCBAA/"+year
    if(os.path.exists(path)):
        pass
    else:
        os.makedirs(path)
    return path
